"""
考勤 Pipeline 运行器 - 封装现有 attendance-pipeline 调用
只产出 Excel 文件（清洗后数据、指标计算后数据、透视分析），不生成 HTML 报告。
"""
import sys
import os
import json
import subprocess
import traceback
from pathlib import Path
from typing import Dict, Optional, Callable, List

from app.config import PIPELINE_DIR
from app.models.schemas import ProcessingStep, SessionStatus

# Pipeline 脚本目录（从 config 读取，支持 Docker 环境）
PIPELINE_SOURCE = Path(PIPELINE_DIR) if PIPELINE_DIR else Path(__file__).resolve().parent.parent.parent / "pipeline"
if str(PIPELINE_SOURCE) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SOURCE))

# Python 解释器（Docker/Linux 中直接用 python，Windows 用管理版本的 venv）
if os.name == "nt":
    # Windows: 使用 venv 里的 python，确保有 pandas/openpyxl 等依赖
    _win_venv = r"C:\Users\Administrator\.workbuddy\binaries\python\envs\default\Scripts\python.exe"
    PYTHON_EXE = _win_venv if os.path.exists(_win_venv) else "python"
else:
    PYTHON_EXE = "python"


# ------ 只保留 Excel 产出步骤 ------
EXCEL_STEPS = [
    ("clean",   ProcessingStep.CLEAN,   10, "清洗后数据.xlsx"),
    ("metrics", ProcessingStep.METRICS, 45, "指标计算后数据.xlsx"),
    ("pivot",   ProcessingStep.PIVOT,   80, "透视分析.xlsx"),
]


class PipelineRunner:
    """
    封装考勤处理全流程，通过 subprocess 调用原始 main.py 的子命令。
    只产出 3 个 Excel 文件，不生成 HTML 报告。
    同时记录处理状态到状态文件。
    """

    def __init__(self, session_id: str, workspace: Path):
        self.session_id = session_id
        self.workspace = workspace
        self.status_file = workspace / "status.json"
        self.progress = 0.0
        self._load_or_init_status()

    def _load_or_init_status(self):
        if self.status_file.exists():
            with open(self.status_file, "r", encoding="utf-8") as f:
                self.state = json.load(f)
        else:
            self.state = {
                "session_id": self.session_id,
                "status": SessionStatus.UPLOADED.value,
                "current_step": None,
                "progress": 0.0,
                "steps_completed": [],
                "error": None,
                "started_at": None,
                "completed_at": None,
            }
            self._save()

    def _save(self):
        with open(self.status_file, "w", encoding="utf-8") as f:
            json.dump(self.state, f, ensure_ascii=False, indent=2)

    def _update(self, step=None, progress=None, status=None, error=None,
                started_at=None, completed_at=None):
        if step is not None:
            self.state["current_step"] = step.value if isinstance(step, ProcessingStep) else step
        if progress is not None:
            self.state["progress"] = progress
        if status is not None:
            self.state["status"] = status.value if isinstance(status, SessionStatus) else status
        if error is not None:
            self.state["error"] = error
        if started_at is not None:
            self.state["started_at"] = started_at
        if completed_at is not None:
            self.state["completed_at"] = completed_at
        self._save()

    def _run_subcmd(self, subcmd, step, progress, extra_args=None):
        """运行 main.py 子命令
        
        命令格式: python main.py --workspace <ws> [--roster-index N] <subcmd>
        注意：--roster-index 等全局参数必须在子命令名称之前
        """
        from datetime import datetime
        self._update(step=step, progress=progress, status=SessionStatus.PROCESSING)

        # 全局参数放在子命令之前
        cmd = [
            PYTHON_EXE,
            str(PIPELINE_SOURCE / "main.py"),
            "--workspace", str(self.workspace),
        ]
        if extra_args:
            cmd.extend(extra_args)  # --roster-index N 等全局参数
        cmd.append(subcmd)  # 子命令放最后

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=600,
                env={
                    **os.environ,
                    "PYTHONIOENCODING": "utf-8",
                    # 确保 pipeline 目录在 PYTHONPATH 中（subprocess 需要能找到 config, pipeline 等模块）
                    "PYTHONPATH": str(PIPELINE_SOURCE) + os.pathsep + os.environ.get("PYTHONPATH", ""),
                },
            )
            if result.returncode != 0:
                error_msg = result.stderr.strip() or result.stdout.strip() or f"返回码: {result.returncode}"
                # 截取最后 2000 字符，避免过长
                if len(error_msg) > 2000:
                    error_msg = "...(省略前部分)\n" + error_msg[-2000:]
                print(f"[Pipeline] 步骤 {step.value} 失败:\nSTDOUT: {result.stdout[-500:] if result.stdout else ''}\nSTDERR: {result.stderr[-1000:] if result.stderr else ''}")
                self._update(error=f"步骤 {step.value} 失败: {error_msg}", status=SessionStatus.FAILED)
                return False

            self.state["steps_completed"].append(step.value)
            self._save()
            return True

        except subprocess.TimeoutExpired:
            self._update(error=f"步骤 {step.value} 超时", status=SessionStatus.FAILED)
            return False
        except Exception as e:
            self._update(error=f"步骤 {step.value} 异常: {str(e)}", status=SessionStatus.FAILED)
            return False

    def run(self, roster_index=None, progress_callback=None):
        """
        执行 Excel 产出流程: clean → metrics → pivot（不产出 HTML）
        """
        from datetime import datetime
        self._update(status=SessionStatus.PROCESSING, started_at=datetime.now().isoformat())

        extra = []
        if roster_index is not None:
            extra = ["--roster-index", str(roster_index)]

        for subcmd, step, progress, filename in EXCEL_STEPS:
            if not self._run_subcmd(subcmd, step, progress, extra):
                return False
            if progress_callback:
                progress_callback(progress, step.value)

        self._update(
            status=SessionStatus.COMPLETED,
            progress=100.0,
            completed_at=datetime.now().isoformat(),
        )
        return True

    def get_status(self):
        self._load_or_init_status()
        return self.state

    # ------ Excel 文件相关 ------

    def get_excel_files(self) -> List[Dict]:
        """列出处理产出的 Excel 文件（供下载）"""
        files = []
        for _, _, _, filename in EXCEL_STEPS:
            fpath = self.workspace / filename
            if fpath.exists():
                files.append({
                    "filename": filename,
                    "label": self._file_label(filename),
                    "size_bytes": fpath.stat().st_size,
                    "download_url": f"/api/v1/process/{self.session_id}/download/{filename}",
                })
        return files

    @staticmethod
    def _file_label(filename: str) -> str:
        labels = {
            "清洗后数据.xlsx": "清洗后数据",
            "指标计算后数据.xlsx": "指标计算后数据",
            "透视分析.xlsx": "透视分析",
        }
        return labels.get(filename, filename)

    def get_summary(self) -> dict:
        """从 Excel 文件中提取处理摘要"""
        summary = {
            "session_id": self.session_id,
            "status": self.state.get("status", "unknown"),
            "error": self.state.get("error"),
        }

        # 尝试从指标文件读取基本统计
        metrics_path = self.workspace / "指标计算后数据.xlsx"
        if metrics_path.exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(metrics_path, read_only=True)
                ws = wb.active
                # 用 openpyxl 数行（含表头）
                row_count = ws.max_row - 1 if ws.max_row else 0  # 减表头
                wb.close()
                summary["total_people"] = row_count
            except Exception:
                pass

        # 尝试从清洗后数据读取
        clean_path = self.workspace / "清洗后数据.xlsx"
        if clean_path.exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(clean_path, read_only=True)
                ws = wb.active
                if ws.max_row:
                    # 读取表头找到排班相关列
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    cols = {str(h): i for i, h in enumerate(headers) if h}

                    scheduled = 0
                    correct = 0
                    over_8h = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # 排班状态
                        schedule_col = cols.get("排班状态")
                        if schedule_col is not None and row[schedule_col] == "是":
                            scheduled += 1
                        # 排班正确
                        correct_col = cols.get("排班正确")
                        if correct_col is not None and row[correct_col] == "是":
                            correct += 1
                        # 日超 8H
                        oh_col = cols.get("日超8H标记")
                        if oh_col is not None and row[oh_col] == "是":
                            over_8h += 1

                    total = row_count = ws.max_row - 1
                    if total > 0:
                        summary["total_people"] = total
                        summary["scheduled_count"] = scheduled
                        summary["scheduled_rate"] = f"{scheduled / total * 100:.1f}%"
                        summary["correct_count"] = correct
                        summary["correct_rate"] = f"{correct / total * 100:.1f}%"
                        summary["over_8h_count"] = over_8h
                wb.close()
            except Exception:
                pass

        return summary
